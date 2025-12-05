import xml.etree.ElementTree as ET
from tkinter import Tk, filedialog
import openpyxl
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
import os

def selecionar_arquivos():
    root = Tk()
    root.withdraw()
    return filedialog.askopenfilenames(
        title="Selecione os arquivos XML da NFe",
        filetypes=[("Arquivos XML", "*.xml"), ("Todos os arquivos", "*.*")]
    )

def extrair_itens_det(caminho_xml):
    try:
        tree = ET.parse(caminho_xml)
        root = tree.getroot()
        ns = {'ns': 'http://www.portalfiscal.inf.br/nfe'}

        inf_nfe = root.find('.//ns:infNFe', ns)
        dados_basicos = {
            'chave': inf_nfe.get('Id')[3:] if inf_nfe is not None else None,
            'numero': root.find('.//ns:nNF', ns).text if root.find('.//ns:nNF', ns) is not None else None,
            'emissao': root.find('.//ns:dhEmi', ns).text if root.find('.//ns:dhEmi', ns) is not None else None,
            'emitente': root.find('.//ns:emit/ns:xNome', ns).text if root.find('.//ns:emit/ns:xNome', ns) is not None else None,
            'destinatario': root.find('.//ns:dest/ns:xNome', ns).text if root.find('.//ns:dest/ns:xNome', ns) is not None else None,
            'total': float(root.find('.//ns:vNF', ns).text) if root.find('.//ns:vNF', ns) is not None else None
        }

        itens = []
        for det in root.findall('.//ns:det', ns):
            prod = det.find('ns:prod', ns)
            imposto = det.find('ns:imposto', ns)
            
            item = {
                'nItem': det.get('nItem'),
                'cProd': prod.find('ns:cProd', ns).text if prod.find('ns:cProd', ns) is not None else None,
                'xProd': prod.find('ns:xProd', ns).text if prod.find('ns:xProd', ns) is not None else None,
                'NCM': prod.find('ns:NCM', ns).text if prod.find('ns:NCM', ns) is not None else None,
                'CFOP': prod.find('ns:CFOP', ns).text if prod.find('ns:CFOP', ns) is not None else None,
                'uCom': prod.find('ns:uCom', ns).text if prod.find('ns:uCom', ns) is not None else None,
                'qCom': float(prod.find('ns:qCom', ns).text) if prod.find('ns:qCom', ns) is not None else None,
                'vUnCom': float(prod.find('ns:vUnCom', ns).text) if prod.find('ns:vUnCom', ns) is not None else None,
                'vProd': float(prod.find('ns:vProd', ns).text) if prod.find('ns:vProd', ns) is not None else None,
                'cEAN': prod.find('ns:cEAN', ns).text if prod.find('ns:cEAN', ns) is not None else None,
            }

            icms = imposto.find('ns:ICMS/ns:ICMS00', ns) or imposto.find('ns:ICMS/ns:ICMS20', ns)
            if icms is not None:
                item.update({
                    'pICMS': float(icms.find('ns:pICMS', ns).text) if icms.find('ns:pICMS', ns) is not None else None,
                    'vICMS': float(icms.find('ns:vICMS', ns).text) if icms.find('ns:vICMS', ns) is not None else None,
                })

            ipi = imposto.find('ns:IPI/ns:IPITrib', ns)
            if ipi is not None:
                item.update({
                    'pIPI': float(ipi.find('ns:pIPI', ns).text) if ipi.find('ns:pIPI', ns) is not None else None,
                    'vIPI': float(ipi.find('ns:vIPI', ns).text) if ipi.find('ns:vIPI', ns) is not None else None,
                })

            itens.append(item)

        return dados_basicos, itens

    except Exception as e:
        print(f"Erro ao processar {caminho_xml}: {str(e)}")
        return None

def exportar_todos_para_excel(saida_excel, dados_completos):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Itens NFe"

    colunas = [
        'Chave NFe', 'Número', 'Emissão', 'Emitente', 'Destinatário', 'Total NFe',
        'Item', 'Código', 'Descrição', 'NCM', 'CFOP', 'Unidade', 'Quantidade',
        'Valor Unitário', 'Valor Total', 'EAN', '% ICMS', 'Valor ICMS', '% IPI', 'Valor IPI'
    ]

    for col_num, cabecalho in enumerate(colunas, 1):
        celula = ws.cell(row=1, column=col_num, value=cabecalho)
        celula.font = Font(bold=True)

    linha = 2
    for dados, itens in dados_completos:
        for item in itens:
            valores = [
                dados['chave'], dados['numero'], dados['emissao'], dados['emitente'],
                dados['destinatario'], dados['total'], item['nItem'], item['cProd'], item['xProd'],
                item['NCM'], item['CFOP'], item['uCom'], item['qCom'], item['vUnCom'],
                item['vProd'], item['cEAN'], item.get('pICMS'), item.get('vICMS'),
                item.get('pIPI'), item.get('vIPI')
            ]
            for col_num, valor in enumerate(valores, 1):
                ws.cell(row=linha, column=col_num, value=valor)
            linha += 1

    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if cell.value and len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        ws.column_dimensions[column].width = (max_length + 2) * 1.2

    wb.save(saida_excel)
    print(f"\nArquivo consolidado salvo como: {saida_excel}")

def main():
    print("=== PROCESSADOR DE NFes EM LOTE ===")
    arquivos = selecionar_arquivos()

    if not arquivos:
        print("Nenhum arquivo selecionado.")
        return

    dados_completos = []
    for arquivo in arquivos:
        print(f"\nProcessando: {arquivo}")
        resultado = extrair_itens_det(arquivo)
        if resultado:
            dados, itens = resultado
            print(f"✔️  {len(itens)} itens extraídos da NFe {dados['chave']}")
            dados_completos.append((dados, itens))

    if dados_completos:
        exportar_todos_para_excel("NFe_Itens_Consolidado.xlsx", dados_completos)
    else:
        print("Nenhuma NFe válida processada.")

if __name__ == "__main__":
    main()
