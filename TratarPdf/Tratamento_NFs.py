import xml.etree.ElementTree as ET
from tkinter import Tk, filedialog
import openpyxl
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
import os
import re

def selecionar_arquivos():
    root = Tk()
    root.withdraw()
    return filedialog.askopenfilenames(
        title="Selecione os arquivos XML da NFe",
        filetypes=[("Arquivos XML", "*.xml"), ("Todos os arquivos", "*.*")]
    )

def extrair_itens_det(caminho_xml):
    try:
        tree = ET.parse(caminho_xml)
        root = tree.getroot()
        ns = {'ns': 'http://www.portalfiscal.inf.br/nfe'}

        inf_nfe = root.find('.//ns:infNFe', ns)
        dados_basicos = {
            'chave': inf_nfe.get('Id')[3:] if inf_nfe is not None else None,
            'numero': root.find('.//ns:nNF', ns).text if root.find('.//ns:nNF', ns) is not None else None,
            'emissao': root.find('.//ns:dhEmi', ns).text if root.find('.//ns:dhEmi', ns) is not None else None,
            'emitente': root.find('.//ns:emit/ns:xNome', ns).text if root.find('.//ns:emit/ns:xNome', ns) is not None else None,
            'destinatario': root.find('.//ns:dest/ns:xNome', ns).text if root.find('.//ns:dest/ns:xNome', ns) is not None else None,
            'total': float(root.find('.//ns:vNF', ns).text) if root.find('.//ns:vNF', ns) is not None else None
        }

        itens = []
        for det in root.findall('.//ns:det', ns):
            prod = det.find('ns:prod', ns)
            imposto = det.find('ns:imposto', ns)
            
            item = {
                'nItem': det.get('nItem'),
                'cProd': prod.find('ns:cProd', ns).text if prod.find('ns:cProd', ns) is not None else None,
                'xProd': prod.find('ns:xProd', ns).text if prod.find('ns:xProd', ns) is not None else None,
                'NCM': prod.find('ns:NCM', ns).text if prod.find('ns:NCM', ns) is not None else None,
                'CFOP': prod.find('ns:CFOP', ns).text if prod.find('ns:CFOP', ns) is not None else None,
                'uCom': prod.find('ns:uCom', ns).text if prod.find('ns:uCom', ns) is not None else None,
                'qCom': float(prod.find('ns:qCom', ns).text) if prod.find('ns:qCom', ns) is not None else None,
                'vUnCom': float(prod.find('ns:vUnCom', ns).text) if prod.find('ns:vUnCom', ns) is not None else None,
                'vProd': float(prod.find('ns:vProd', ns).text) if prod.find('ns:vProd', ns) is not None else None,
                'cEAN': prod.find('ns:cEAN', ns).text if prod.find('ns:cEAN', ns) is not None else None,
            }

            # Busca dinâmica por todos os campos de ICMS
            if imposto is not None:
                icms = imposto.find('ns:ICMS', ns)
                if icms is not None:
                    for icms_tipo in icms:
                        # Remove o namespace do nome do campo
                        campo = icms_tipo.tag.split('}')[-1] if '}' in icms_tipo.tag else icms_tipo.tag
                        for campo_icms in icms_tipo:
                            # Remove o namespace e adiciona o prefixo ICMS_
                            nome_campo = "ICMS_" + (campo_icms.tag.split('}')[-1] if '}' in campo_icms.tag else campo_icms.tag)
                            try:
                                valor = float(campo_icms.text) if campo_icms.text else None
                            except (ValueError, TypeError):
                                valor = campo_icms.text
                            item[nome_campo] = valor

            # IPI (mantido como estava)
            ipi = imposto.find('ns:IPI/ns:IPITrib', ns) if imposto is not None else None
            if ipi is not None:
                item.update({
                    'IPI_pIPI': float(ipi.find('ns:pIPI', ns).text) if ipi.find('ns:pIPI', ns) is not None else None,
                    'IPI_vIPI': float(ipi.find('ns:vIPI', ns).text) if ipi.find('ns:vIPI', ns) is not None else None,
                })

            itens.append(item)

        return dados_basicos, itens

    except Exception as e:
        print(f"Erro ao processar {caminho_xml}: {str(e)}")
        return None
def exportar_todos_para_excel(saida_excel, dados_completos):
    if not dados_completos:
        print("Nenhum dado para exportar.")
        return

    # Primeiro passamos por todos os itens para descobrir todas as colunas ICMS existentes
    todas_colunas = {
        'Chave NFe', 'Número', 'Emissão', 'Emitente', 'Destinatário', 'Total NFe',
        'Item', 'Código', 'Descrição', 'NCM', 'CFOP', 'Unidade', 'Quantidade',
        'Valor Unitário', 'Valor Total', 'EAN'
    }
    
    # Adiciona colunas de IPI fixas
    todas_colunas.update(['IPI_pIPI', 'IPI_vIPI'])
    
    # Coleta todas as colunas ICMS dinâmicas
    colunas_icms = set()
    for dados, itens in dados_completos:
        for item in itens:
            for chave in item.keys():
                if chave.startswith('ICMS_'):
                    colunas_icms.add(chave)
    
    # Ordena as colunas ICMS para manter consistência
    colunas_icms_ordenadas = sorted(colunas_icms)
    todas_colunas.update(colunas_icms_ordenadas)
    
    # Cria a lista ordenada de colunas
    colunas_ordenadas = [
        'Chave NFe', 'Número', 'Emissão', 'Emitente', 'Destinatário', 'Total NFe',
        'Item', 'Código', 'Descrição', 'NCM', 'CFOP', 'Unidade', 'Quantidade',
        'Valor Unitário', 'Valor Total', 'EAN'
    ]
    
    # Adiciona as colunas ICMS ordenadas
    colunas_ordenadas.extend(colunas_icms_ordenadas)
    
    # Adiciona as colunas IPI no final
    colunas_ordenadas.extend(['IPI_pIPI', 'IPI_vIPI'])

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Itens NFe"

    # Escreve os cabeçalhos
    for col_num, cabecalho in enumerate(colunas_ordenadas, 1):
        celula = ws.cell(row=1, column=col_num, value=cabecalho)
        celula.font = Font(bold=True)

    # Escreve os dados
    linha = 2
    for dados, itens in dados_completos:
        for item in itens:
            valores = [
                dados['chave'], dados['numero'], dados['emissao'], dados['emitente'],
                dados['destinatario'], dados['total'], item['nItem'], item['cProd'], 
                item['xProd'], item['NCM'], item['CFOP'], item['uCom'], item['qCom'], 
                item['vUnCom'], item['vProd'], item['cEAN']
            ]
            
            # Adiciona os valores ICMS dinâmicos
            for coluna_icms in colunas_icms_ordenadas:
                valores.append(item.get(coluna_icms))
            
            # Adiciona os valores IPI
            valores.extend([
                item.get('IPI_pIPI'),
                item.get('IPI_vIPI')
            ])
            
            # Escreve na planilha
            for col_num, valor in enumerate(valores, 1):
                ws.cell(row=linha, column=col_num, value=valor)
            linha += 1

    # Ajusta a largura das colunas
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if cell.value and len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2) * 1.2
        ws.column_dimensions[column].width = adjusted_width if adjusted_width > 10 else 10

    wb.save(saida_excel)
    print(f"\nArquivo consolidado salvo como: {saida_excel}")

def main():
    print("=== PROCESSADOR DE NFes EM LOTE ===")
    arquivos = selecionar_arquivos()

    if not arquivos:
        print("Nenhum arquivo selecionado.")
        return

    dados_completos = []
    for arquivo in arquivos:
        print(f"\nProcessando: {arquivo}")
        resultado = extrair_itens_det(arquivo)
        if resultado:
            dados, itens = resultado
            print(f"✔️  {len(itens)} itens extraídos da NFe {dados['chave']}")
            dados_completos.append((dados, itens))

    if dados_completos:
        exportar_todos_para_excel("NFe_Itens_Consolidado.xlsx", dados_completos)
    else:
        print("Nenhuma NFe válida processada.")

if __name__ == "__main__":
    main()